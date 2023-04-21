import openpyxl

def load_data():

    inventory_file = openpyxl.load_workbook("./input/inventory.xlsx")
    raw_data = inventory_file["DataSheet"]
    return raw_data

def products_per_supplier():
    
    product_list = load_data()

    products_per_supplier = {}
    
    for product_row in range(2, product_list.max_row + 1):

        supplier_name = product_list.cell(product_row, 4).value

        if supplier_name in products_per_supplier:
            current_num_products = products_per_supplier[supplier_name]
            products_per_supplier[supplier_name] = current_num_products + 1

        else:
            products_per_supplier[supplier_name] = 1

    return products_per_supplier